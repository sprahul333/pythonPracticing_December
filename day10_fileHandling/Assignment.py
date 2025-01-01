
#Write Data to the File:
#Write a program to write data to the file.

f2=open("assignment.txt","w")
f2.write("Selenium is for Web Automation, Selenium is free and open source and Selenium is free. Selenium Grid")
f2.close()

#Read the Data from the File:
f2=open("assignment.txt","r")
data=f2.read()

#Read the duplicates the words in the file
words=data.split(" ")

wordCount={}

for word in words:
    if word in wordCount:
        wordCount[word]=wordCount[word]+1
    else:
        wordCount[word]=1

#Print the Duplicate words in the map

for key in wordCount:
    if wordCount[key]>1:
        print(key,wordCount[key])

#Write the data to the excel file

import openpyxl

# Create a new workbook and select the active worksheet
wb = openpyxl.Workbook() # Create a new workbook
ws = wb.active # Select the active worksheet
ws.title = "MySheet"  # Naming the sheet

#Print the duplicate words in the excel file
ws.cell(1, 1, "Word") # Write the word to the first column
ws.cell(1, 2, "Count") # Write the count to the second column

row = 2
for key in wordCount:
    if wordCount[key] > 1:
        ws.cell(row, 1, key) # Write the key to the first column
        ws.cell(row, 2, wordCount[key]) # Write the value to the second column
        row += 1

#Get Total Number of Rows and Columns
print(ws.max_row)
print(ws.max_column)

#Get the value of the cell
print(ws.cell(2,1).value)

#Get the value of the cell
print(ws.cell(2,2).value)

# Save the workbook
wb.save("assignment.xlsx")

#Close the file
f2.close()


